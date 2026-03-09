"""
Seed the database with all credit cards parsed from Financial.xlsx.

Run directly (from the backend/ directory):
    python -m app.seed_data

Or from the project root:
    cd backend && python -m app.seed_data

Seeding order (respects FK dependencies):
    1. Issuers
    2. Currencies          (FK -> issuers)
    3. EcosystemBoosts     (FK -> issuers, currencies)
    4. EcosystemBoostAnchors after cards are created
    5. Cards               (FK -> issuers, currencies, ecosystem_boosts)
    6. EcosystemBoostAnchors (FK -> ecosystem_boosts, cards)
    7. CardCategoryMultipliers / CardCredits / SpendCategories
"""

from __future__ import annotations

import asyncio
import os
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.database import AsyncSessionLocal, create_tables
from app.models import (
    Card,
    CardCategoryMultiplier,
    CardCredit,
    Currency,
    EcosystemBoost,
    EcosystemBoostAnchor,
    Issuer,
    SpendCategory,
)

XLSX_PATH = Path(__file__).resolve().parent.parent.parent / "docs" / "Financial.xlsx"

CATEGORIES = [
    "All Other",
    "Dining",
    "Groceries",
    "Personal Care",
    "Drugstores",
    "Fitness",
    "Gas",
    "Rideshare",
    "Transit",
    "Entertainment",
    "Streaming",
    "Software, Hardware",
    "Internet",
    "Phone",
    "Airlines",
    "Hotels",
    "Rotating",
]

CREDIT_TYPES = [
    "Misc. Travel",
    "Hotel Collection",
    "Hotel Status",
    "Rental Car Status",
    "Lounge Access",
    "Airport Security",
    "Flight Credits",
    "Status Headstart",
    "Dining",
    "Streaming",
    "Rideshare & Delivery",
    "Live Entertainment",
    "Miscellaneous",
]

# ---------------------------------------------------------------------------
# Issuer seed data
# ---------------------------------------------------------------------------

ISSUERS = [
    "American Express",
    "Chase",
    "Capital One",
    "Citi",
    "Bilt",
    "Delta / American Express",
    "Hilton / American Express",
]

# ---------------------------------------------------------------------------
# Currency seed data
# Each entry: (issuer_name, currency_name, cpp, is_cashback, is_transferable, comparison_factor)
# ---------------------------------------------------------------------------

CURRENCIES: list[tuple[str, str, float, bool, bool, float]] = [
    # American Express
    ("American Express", "Amex MR",        2.0,  False, True,  1.0),
    # Chase — cashback variant (Freedom Unlimited/Flex alone) + transferable variant
    ("Chase",            "Chase UR Cash",   1.0,  True,  False, 1.0),
    ("Chase",            "Chase UR",        1.5,  False, True,  1.0),
    # Capital One
    ("Capital One",      "Capital One Miles", 1.5, False, True, 1.0),
    # Citi — cashback variant + transferable variant
    ("Citi",             "Citi TY Cash",    1.0,  True,  False, 1.0),
    ("Citi",             "Citi TY",         1.5,  False, True,  1.0),
    # Bilt
    ("Bilt",             "Bilt Rewards",    1.5,  False, True,  1.0),
    # Delta cobrand — non-transferable, comparison_factor < 1 to normalise vs MR/UR
    ("Delta / American Express", "Delta SkyMiles", 1.2, False, False, 0.85),
    # Hilton cobrand — non-transferable, low cpp but high earn multipliers
    ("Hilton / American Express", "Hilton Honors", 0.5, False, False, 1.0),
]

# ---------------------------------------------------------------------------
# Ecosystem boost seed data
# Each entry: (issuer_name, boost_name, boosted_currency_name, description)
# ---------------------------------------------------------------------------

ECOSYSTEM_BOOSTS: list[tuple[str, str, str, str]] = [
    (
        "Chase",
        "Chase UR Upgrade",
        "Chase UR",
        (
            "When a Chase Sapphire Reserve, Sapphire Preferred, or Ink Preferred "
            "is in the wallet, cashback-earning Chase cards (Freedom Unlimited, "
            "Freedom Flex) convert to transferable Chase UR points."
        ),
    ),
    (
        "Citi",
        "Citi TY Upgrade",
        "Citi TY",
        (
            "When a Citi Strata Elite is in the wallet, other Citi ThankYou cards "
            "(Strata Premier, Custom Cash, Double Cash) convert from cashback to "
            "transferable Citi TY points."
        ),
    ),
]

# ---------------------------------------------------------------------------
# Card -> issuer / currency / ecosystem_boost mappings
# ---------------------------------------------------------------------------

# card name -> issuer name
CARD_ISSUER_MAP: dict[str, str] = {
    "American Express Platinum":        "American Express",
    "American Express Gold":            "American Express",
    "American Express Business Gold":   "American Express",
    "American Express Blue Business Plus": "American Express",
    "Chase Sapphire Reserve":           "Chase",
    "Chase Sapphire Preferred":         "Chase",
    "Chase Ink Preferred":              "Chase",
    "Chase Ink Cash":                   "Chase",
    "Chase Freedom Unlimited":          "Chase",
    "Chase Freedom Flex":               "Chase",
    "Capital One Venture X":            "Capital One",
    "Capital One Savor":                "Capital One",
    "Citi Strata Elite":                "Citi",
    "Citi Strata Premier":              "Citi",
    "Citi Strata":                      "Citi",
    "Citi Custom Cash":                 "Citi",
    "Citi Double Cash":                 "Citi",
    "Bilt Palladium":                   "Bilt",
    "Bilt Obsidian":                    "Bilt",
    "Bilt Blue":                        "Bilt",
    "Delta SkyMiles Gold":              "Delta / American Express",
    "Delta SkyMiles Gold Business":     "Delta / American Express",
    "Delta SkyMiles Platinum":          "Delta / American Express",
    "Delta SkyMiles Reserve":           "Delta / American Express",
    "Hilton Honors Surpass":            "Hilton / American Express",
    "Hilton Honors Aspire":             "Hilton / American Express",
}

# card name -> default currency name
# Cashback-capable Chase/Citi cards default to their cashback currency;
# the ecosystem boost will upgrade them at runtime.
CARD_CURRENCY_MAP: dict[str, str] = {
    "American Express Platinum":        "Amex MR",
    "American Express Gold":            "Amex MR",
    "American Express Business Gold":   "Amex MR",
    "American Express Blue Business Plus": "Amex MR",
    "Chase Sapphire Reserve":           "Chase UR",
    "Chase Sapphire Preferred":         "Chase UR",
    "Chase Ink Preferred":              "Chase UR",
    "Chase Ink Cash":                   "Chase UR",
    "Chase Freedom Unlimited":          "Chase UR Cash",
    "Chase Freedom Flex":               "Chase UR Cash",
    "Capital One Venture X":            "Capital One Miles",
    "Capital One Savor":                "Capital One Miles",
    "Citi Strata Elite":                "Citi TY",
    "Citi Strata Premier":              "Citi TY Cash",
    "Citi Strata":                      "Citi TY Cash",
    "Citi Custom Cash":                 "Citi TY Cash",
    "Citi Double Cash":                 "Citi TY Cash",
    "Bilt Palladium":                   "Bilt Rewards",
    "Bilt Obsidian":                    "Bilt Rewards",
    "Bilt Blue":                        "Bilt Rewards",
    "Delta SkyMiles Gold":              "Delta SkyMiles",
    "Delta SkyMiles Gold Business":     "Delta SkyMiles",
    "Delta SkyMiles Platinum":          "Delta SkyMiles",
    "Delta SkyMiles Reserve":           "Delta SkyMiles",
    "Hilton Honors Surpass":            "Hilton Honors",
    "Hilton Honors Aspire":             "Hilton Honors",
}

# card name -> ecosystem boost name (only cards that BENEFIT from a boost)
CARD_BOOST_MAP: dict[str, str] = {
    "Chase Freedom Unlimited":  "Chase UR Upgrade",
    "Chase Freedom Flex":       "Chase UR Upgrade",
    "Citi Strata Premier":      "Citi TY Upgrade",
    "Citi Strata":              "Citi TY Upgrade",
    "Citi Custom Cash":         "Citi TY Upgrade",
    "Citi Double Cash":         "Citi TY Upgrade",
}

# boost name -> list of anchor card names
BOOST_ANCHORS: dict[str, list[str]] = {
    "Chase UR Upgrade": [
        "Chase Sapphire Reserve",
        "Chase Sapphire Preferred",
        "Chase Ink Preferred",
    ],
    "Citi TY Upgrade": [
        "Citi Strata Elite",
    ],
}


# ---------------------------------------------------------------------------
# XLSX parsing
# ---------------------------------------------------------------------------


def parse_xlsx() -> dict:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    ws = wb["Credit Card Tool"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    card_cols = list(range(5, 57, 2))
    card_names = [rows[0][c] for c in card_cols]

    all_cards: dict = {}
    for col, name in zip(card_cols, card_names):
        if not name:
            continue
        mult_col = col + 1

        annual_fee        = rows[6][col] or 0
        sub_points        = rows[7][col] or 0
        sub_min_spend     = rows[9][col]
        sub_months        = rows[10][col]
        sub_spend_points  = rows[13][col] or 0
        annual_bonus_pts  = rows[8][col] or 0

        multipliers: dict[str, float] = {}
        for cat, row_idx in zip(CATEGORIES, range(18, 35)):
            val = rows[row_idx][mult_col]
            multipliers[cat] = float(val) if val is not None else 1.0

        credits: dict[str, dict] = {}
        for ctype, row_idx in zip(CREDIT_TYPES, range(35, 48)):
            label = rows[row_idx][col]
            value = rows[row_idx][mult_col]
            if label is not None or (value is not None and value != 0):
                credits[ctype] = {
                    "label": str(label) if label else "",
                    "value": float(value) if value is not None else 0.0,
                }

        all_cards[name] = {
            "name":                name,
            "annual_fee":          float(annual_fee),
            "sub_points":          int(sub_points),
            "sub_min_spend":       int(sub_min_spend) if sub_min_spend is not None else None,
            "sub_months":          int(sub_months) if sub_months is not None else None,
            "sub_spend_points":    int(sub_spend_points),
            "annual_bonus_points": int(annual_bonus_pts),
            "multipliers":         multipliers,
            "credits":             credits,
        }

    return all_cards


def parse_spend_categories() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    ws = wb["Credit Card Tool"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    result = []
    for cat, row_idx in zip(CATEGORIES, range(18, 35)):
        spend = rows[row_idx][4]
        result.append({"category": cat, "annual_spend": float(spend) if spend else 0.0})
    return result


# ---------------------------------------------------------------------------
# Seed helpers
# ---------------------------------------------------------------------------


async def _upsert_issuer(session: AsyncSession, name: str) -> Issuer:
    existing = await session.execute(select(Issuer).where(Issuer.name == name))
    obj = existing.scalar_one_or_none()
    if obj is None:
        obj = Issuer(name=name)
        session.add(obj)
        await session.flush()
    return obj


async def _upsert_currency(
    session: AsyncSession,
    issuer: Issuer,
    name: str,
    cpp: float,
    is_cashback: bool,
    is_transferable: bool,
    comparison_factor: float,
) -> Currency:
    existing = await session.execute(select(Currency).where(Currency.name == name))
    obj = existing.scalar_one_or_none()
    if obj is None:
        obj = Currency(
            issuer_id=issuer.id,
            name=name,
            cents_per_point=cpp,
            is_cashback=is_cashback,
            is_transferable=is_transferable,
            comparison_factor=comparison_factor,
        )
        session.add(obj)
        await session.flush()
    else:
        obj.issuer_id = issuer.id
        obj.cents_per_point = cpp
        obj.is_cashback = is_cashback
        obj.is_transferable = is_transferable
        obj.comparison_factor = comparison_factor
        await session.flush()
    return obj


async def _upsert_boost(
    session: AsyncSession,
    issuer: Issuer,
    name: str,
    boosted_currency: Currency,
    description: str,
) -> EcosystemBoost:
    existing = await session.execute(
        select(EcosystemBoost).where(EcosystemBoost.name == name)
    )
    obj = existing.scalar_one_or_none()
    if obj is None:
        obj = EcosystemBoost(
            issuer_id=issuer.id,
            boosted_currency_id=boosted_currency.id,
            name=name,
            description=description,
        )
        session.add(obj)
        await session.flush()
    else:
        obj.issuer_id = issuer.id
        obj.boosted_currency_id = boosted_currency.id
        obj.description = description
        await session.flush()
    return obj


# ---------------------------------------------------------------------------
# Main seed function
# ---------------------------------------------------------------------------


async def seed(session: AsyncSession) -> None:
    """Upsert all reference data and cards."""

    # 1. Issuers
    issuer_objs: dict[str, Issuer] = {}
    for name in ISSUERS:
        issuer_objs[name] = await _upsert_issuer(session, name)
    print(f"  Issuers:  {len(issuer_objs)}")

    # 2. Currencies
    currency_objs: dict[str, Currency] = {}
    for issuer_name, cur_name, cpp, is_cb, is_tf, cf in CURRENCIES:
        currency_objs[cur_name] = await _upsert_currency(
            session,
            issuer_objs[issuer_name],
            cur_name,
            cpp,
            is_cb,
            is_tf,
            cf,
        )
    print(f"  Currencies: {len(currency_objs)}")

    # 3. Ecosystem boosts (anchor links come after cards are created)
    boost_objs: dict[str, EcosystemBoost] = {}
    for issuer_name, boost_name, boosted_cur_name, desc in ECOSYSTEM_BOOSTS:
        boost_objs[boost_name] = await _upsert_boost(
            session,
            issuer_objs[issuer_name],
            boost_name,
            currency_objs[boosted_cur_name],
            desc,
        )
    print(f"  EcosystemBoosts: {len(boost_objs)}")

    # 4. Spend categories
    spend_cats = parse_spend_categories()
    for sc in spend_cats:
        existing = await session.execute(
            select(SpendCategory).where(SpendCategory.category == sc["category"])
        )
        obj = existing.scalar_one_or_none()
        if obj is None:
            session.add(SpendCategory(**sc))
        else:
            obj.annual_spend = sc["annual_spend"]
    print(f"  SpendCategories: {len(spend_cats)}")

    # 5. Cards
    all_cards_raw = parse_xlsx()
    card_objs: dict[str, Card] = {}

    for card_name, card_data in all_cards_raw.items():
        multipliers = card_data.pop("multipliers")
        credits_data = card_data.pop("credits")

        issuer_name = CARD_ISSUER_MAP.get(card_name, "Unknown")
        currency_name = CARD_CURRENCY_MAP.get(card_name, "Chase UR Cash")
        boost_name = CARD_BOOST_MAP.get(card_name)

        issuer = issuer_objs.get(issuer_name)
        if issuer is None:
            # Dynamically create unknown issuers
            issuer = await _upsert_issuer(session, issuer_name)
            issuer_objs[issuer_name] = issuer

        currency = currency_objs.get(currency_name)
        if currency is None:
            raise ValueError(
                f"Currency '{currency_name}' not found for card '{card_name}'. "
                "Check CARD_CURRENCY_MAP and CURRENCIES."
            )

        boost = boost_objs.get(boost_name) if boost_name else None

        existing = await session.execute(
            select(Card).where(Card.name == card_data["name"])
        )
        card = existing.scalar_one_or_none()

        if card is None:
            card = Card(
                issuer_id=issuer.id,
                currency_id=currency.id,
                ecosystem_boost_id=boost.id if boost else None,
                **card_data,
            )
            session.add(card)
            await session.flush()
        else:
            card.issuer_id = issuer.id
            card.currency_id = currency.id
            card.ecosystem_boost_id = boost.id if boost else None
            for k, v in card_data.items():
                setattr(card, k, v)
            await session.flush()

        card_objs[card_name] = card

        # Refresh multipliers and credits
        await session.execute(
            CardCategoryMultiplier.__table__.delete().where(
                CardCategoryMultiplier.card_id == card.id
            )
        )
        await session.execute(
            CardCredit.__table__.delete().where(CardCredit.card_id == card.id)
        )
        for cat, mult in multipliers.items():
            session.add(CardCategoryMultiplier(card_id=card.id, category=cat, multiplier=mult))
        for ctype, cinfo in credits_data.items():
            session.add(
                CardCredit(
                    card_id=card.id,
                    credit_name=f"{ctype}: {cinfo['label']}".strip(": "),
                    credit_value=cinfo["value"],
                )
            )

    print(f"  Cards: {len(card_objs)}")

    # 6. Ecosystem boost anchors (requires cards to exist)
    for boost_name, anchor_card_names in BOOST_ANCHORS.items():
        boost = boost_objs[boost_name]
        for anchor_name in anchor_card_names:
            anchor_card = card_objs.get(anchor_name)
            if anchor_card is None:
                print(f"  WARNING: anchor card '{anchor_name}' not found for boost '{boost_name}'")
                continue
            existing = await session.execute(
                select(EcosystemBoostAnchor).where(
                    EcosystemBoostAnchor.boost_id == boost.id,
                    EcosystemBoostAnchor.card_id == anchor_card.id,
                )
            )
            if existing.scalar_one_or_none() is None:
                session.add(
                    EcosystemBoostAnchor(boost_id=boost.id, card_id=anchor_card.id)
                )
    await session.flush()
    print(f"  EcosystemBoostAnchors: seeded")

    await session.commit()
    print("Seed complete.")


async def main() -> None:
    await create_tables()
    async with AsyncSessionLocal() as session:
        await seed(session)


if __name__ == "__main__":
    asyncio.run(main())
